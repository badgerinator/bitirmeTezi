"""
AHP Pairwise Comparison Matrisi Oluşturucu (Formüllü + Başlık Sorunsuz + Üst Üçgen 5)
===================================================================================

6 uzmanın doldurması için AHP matrisleri üretir:
- Alt üçgen hücreler otomatik formül
- Üst üçgen hücrelere başlangıçta 5 yazılır
- Header ve index hizası düzgün
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# Kriter isimleri
criteria = [
    "Deneyim Seviyesi (Kategori)",
    "Yabancı Dil Skoru",
    "Eğitim Seviyesi Skoru",
    "Basic Computer Skills Skoru",
    "Katıldığınız Kurs/Seminer/Sertifika Skoru",
    "Sosyal Aktivite Skoru (0-100)"
]

# Dosya yolu
output_path = "./outputs/ahp_matrices_formullu.xlsx"

# Boş DataFrame oluştur
def create_initial_ahp_matrix(criteria):
    size = len(criteria)
    df_matrix = pd.DataFrame('', index=criteria, columns=criteria)
    for i in range(size):
        for j in range(size):
            if i == j:
                df_matrix.iat[i, j] = 1  # diagonal → 1
            elif i < j:
                df_matrix.iat[i, j] = 5  # üst üçgen → başlangıçta 5
            else:
                df_matrix.iat[i, j] = ''  # alt üçgen → boş (formül eklenecek)
    return df_matrix

# ExcelWriter (openpyxl backend kullanacağız → formül destekli)
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for expert_id in range(1, 7):
        df_matrix = create_initial_ahp_matrix(criteria)
        sheet_name = f"Uzman_{expert_id}"
        # startrow=0, startcol=0 → hizalama düzgün olur
        df_matrix.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0)

# openpyxl ile tekrar aç → formülleri ekle
wb = openpyxl.load_workbook(output_path)

for expert_id in range(1, 7):
    sheet_name = f"Uzman_{expert_id}"
    ws = wb[sheet_name]
    
    n = len(criteria)
    start_row = 2  # 1-based index: 1.row → header, 2.row → veri başlar
    start_col = 2  # 1.col → index, 2.col → veri başlar
    
    # Alt üçgen hücreler için formül yaz
    for i in range(n):
        for j in range(i):
            row = start_row + i
            col = start_col + j
            
            # Üst üçgendeki referans hücresini bul
            ref_row = start_row + j
            ref_col_letter = get_column_letter(start_col + i)
            ref_cell = f"{ref_col_letter}{ref_row}"
            
            # Formül ekle
            ws.cell(row=row, column=col).value = f"=1/{ref_cell}"

# Dosyayı kaydet
wb.save(output_path)

print(f"Formüllü AHP matrisleri başarıyla oluşturuldu ve kaydedildi: {output_path}")
